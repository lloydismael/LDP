"""Versioned Excel template, validation, preview, and atomic import services."""

from __future__ import annotations

import hashlib
import io
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ldp_core.models import (
    Activity,
    ImportJob,
    ImportRowResult,
    LeadershipAward,
    Person,
    School,
    User,
)

TEMPLATE_VERSION = '1.0'
MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_ROWS_PER_SHEET = 10_000
SHEETS = {
    'Schools': [
        'school_id', 'name', 'school_type', 'category', 'address', 'location',
        'district', 'division', 'province', 'region', 'email', 'phone', 'website',
        'founded_year', 'is_active', 'principal_username',
    ],
    'Users': [
        'username', 'first_name', 'last_name', 'email', 'role', 'is_active',
        'person_type', 'school_id', 'contact_number', 'address', 'bio', 'student_id',
        'year_level', 'course_program', 'section', 'scholarship_type',
        'year_started', 'year_ended',
    ],
    'Activities': ['school_id', 'name', 'date', 'description', 'is_approved'],
    'ActivityParticipants': ['school_id', 'activity_name', 'activity_date', 'username'],
    'LeadershipAwards': [
        'recipient_username', 'award_title', 'award_level', 'year_awarded',
        'awarding_body', 'description', 'school_id',
    ],
}
REQUIRED = {
    'Schools': {'name'},
    'Users': {'username', 'first_name', 'last_name', 'role', 'person_type'},
    'Activities': {'school_id', 'name', 'date'},
    'ActivityParticipants': {'school_id', 'activity_name', 'activity_date', 'username'},
    'LeadershipAwards': {'recipient_username', 'award_title', 'award_level', 'year_awarded'},
}
ROLE_FOR_PERSON = {
    Person.Type.PRINCIPAL: User.Role.PRINCIPAL,
    Person.Type.SCHOLAR: User.Role.SCHOLAR,
    Person.Type.PROFESSIONAL: User.Role.PROFESSIONAL,
    Person.Type.STUDENT: User.Role.VIEWER,
    Person.Type.COLLEGE: User.Role.VIEWER,
}


class ImportValidationError(ValueError):
    pass


@dataclass
class RowPlan:
    sheet: str
    row_number: int
    source: dict[str, Any]
    key: str = ''
    action: str = ImportRowResult.Action.UNCHANGED
    changes: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


@dataclass
class ImportPlan:
    rows: list[RowPlan]

    @property
    def errors(self):
        return [row for row in self.rows if row.errors]

    def count(self, action):
        return sum(row.action == action for row in self.rows)


def normalize(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip())


def normalized_key(value: Any) -> str:
    return normalize(value).casefold()


def parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or normalize(value) == '':
        return default
    if isinstance(value, bool):
        return value
    text = normalized_key(value)
    if text in {'true', 'yes', 'y', '1', 'active', 'approved'}:
        return True
    if text in {'false', 'no', 'n', '0', 'inactive', 'unapproved'}:
        return False
    raise ImportValidationError(f'Invalid boolean value: {value!r}.')


def parse_date(value: Any, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize(value)
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ImportValidationError(f'{field_name} must use YYYY-MM-DD.') from exc


def parse_year(value: Any, field_name: str) -> str:
    text = normalize(value)
    if not re.fullmatch(r'\d{4}', text):
        raise ImportValidationError(f'{field_name} must be a four-digit year.')
    year = int(text)
    if year < 1900 or year > date.today().year + 5:
        raise ImportValidationError(f'{field_name} is outside the supported range.')
    return text


def safe_json_value(value: Any):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def json_safe(value: Any):
    if isinstance(value, dict):
        return {key: json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]
    return safe_json_value(value)


def school_key(row: dict[str, Any]) -> str:
    identifier = normalized_key(row.get('school_id'))
    if identifier:
        return identifier
    return '|'.join([
        normalized_key(row.get('name')),
        normalized_key(row.get('division')),
        normalized_key(row.get('location')),
    ])


def activity_key(school_identifier: Any, name: Any, activity_date: Any) -> str:
    date_value = activity_date.isoformat() if isinstance(activity_date, date) else normalize(activity_date)
    return f'{normalized_key(school_identifier)}|{normalized_key(name)}|{date_value}'


def award_key(username: Any, title: Any, year: Any, awarding_body: Any) -> str:
    return '|'.join(map(normalized_key, [username, title, year, awarding_body]))


def workbook_bytes() -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = 'Instructions'
    instructions.append(['LDP Administrative Import Template', f'Version {TEMPLATE_VERSION}'])
    instructions.append(['Workflow', 'Complete the sheets, upload, review every preview result, then confirm.'])
    instructions.append(['Dates', 'Use YYYY-MM-DD. Years use YYYY. Do not use formulas.'])
    instructions.append(['Matching', 'school_id and username are stable external identifiers.'])
    instructions.append(['Passwords', 'Passwords are never imported. New accounts require activation.'])
    instructions['A1'].font = Font(bold=True, size=14)

    enum_options = {
        'school_type': [choice for choice, _ in School.SchoolType.choices],
        'role': [choice for choice, _ in User.Role.choices],
        'person_type': [choice for choice, _ in Person.Type.choices],
        'award_level': [choice for choice, _ in LeadershipAward.AwardLevel.choices],
        'boolean': ['TRUE', 'FALSE'],
    }
    for sheet_name, headers in SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = f'A1:{sheet.cell(1, len(headers)).coordinate}'
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4F46E5')
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(14, min(28, len(header) + 3))
            values = None
            if header in enum_options:
                values = enum_options[header]
            elif header in {'is_active', 'is_approved'}:
                values = enum_options['boolean']
            if values:
                validation = DataValidation(type='list', formula1='"' + ','.join(values) + '"')
                validation.error = 'Select a value from the list.'
                validation.errorTitle = 'Invalid value'
                validation.add(f'{sheet.cell(2, index).coordinate}:{sheet.cell(MAX_ROWS_PER_SHEET + 1, index).coordinate}')
                sheet.add_data_validation(validation)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def read_workbook(content: bytes) -> dict[str, list[tuple[int, dict[str, Any]]]]:
    if len(content) > MAX_IMPORT_BYTES:
        raise ImportValidationError('Workbook exceeds the 10 MB upload limit.')
    if not content.startswith(b'PK'):
        raise ImportValidationError('The uploaded file is not a valid .xlsx workbook.')
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise ImportValidationError('The uploaded workbook could not be opened.') from exc
    missing = [sheet for sheet in SHEETS if sheet not in workbook.sheetnames]
    if missing:
        raise ImportValidationError(f"Missing required sheet(s): {', '.join(missing)}.")

    result = {}
    for sheet_name, expected_headers in SHEETS.items():
        sheet = workbook[sheet_name]
        if sheet.max_row > MAX_ROWS_PER_SHEET + 1:
            raise ImportValidationError(f'{sheet_name} exceeds {MAX_ROWS_PER_SHEET:,} data rows.')
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [normalize(value) for value in first_row]
        missing_headers = [header for header in expected_headers if header not in headers]
        if missing_headers:
            raise ImportValidationError(
                f"{sheet_name} is missing column(s): {', '.join(missing_headers)}."
            )
        rows = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            raw_values = []
            for cell in values:
                if cell.data_type == 'f':
                    raise ImportValidationError(f'Formulas are not allowed ({sheet_name}!{cell.coordinate}).')
                raw_values.append(cell.value)
            source = {header: raw_values[headers.index(header)] for header in expected_headers}
            if not any(normalize(value) for value in source.values()):
                continue
            rows.append((row_number, source))
        result[sheet_name] = rows
    return result


def _choice(value: Any, choices, field_name: str, required=True) -> str:
    text = normalize(value).upper()
    allowed = {choice for choice, _ in choices}
    if not text and not required:
        return ''
    if text not in allowed:
        raise ImportValidationError(f'{field_name} must be one of: {", ".join(sorted(allowed))}.')
    return text


def _changes(instance, proposed: dict[str, Any]) -> dict[str, Any]:
    changes = {}
    for name, value in proposed.items():
        current = getattr(instance, name)
        current_id = current.pk if hasattr(current, 'pk') else current
        value_id = value.pk if hasattr(value, 'pk') else value
        if current_id != value_id:
            changes[name] = {'from': safe_json_value(current_id), 'to': safe_json_value(value_id)}
    return changes


def _find_school(row: dict[str, Any]):
    identifier = normalize(row.get('school_id'))
    if identifier:
        matches = School.objects.filter(school_id__iexact=identifier)
    else:
        matches = School.objects.filter(
            name__iexact=normalize(row.get('name')),
            division__iexact=normalize(row.get('division')),
            location__iexact=normalize(row.get('location')) or 'Philippines',
        )
    if matches.count() > 1:
        raise ImportValidationError('The school key matches multiple database records.')
    return matches.first()


def build_plan(content: bytes) -> ImportPlan:
    parsed = read_workbook(content)
    plans: list[RowPlan] = []
    staged_schools: dict[str, dict[str, Any]] = {}
    staged_users: dict[str, dict[str, Any]] = {}
    staged_activities: dict[str, dict[str, Any]] = {}

    duplicate_keys = {}
    key_functions = {
        'Schools': lambda row: school_key(row),
        'Users': lambda row: normalized_key(row.get('username')),
        'Activities': lambda row: activity_key(row.get('school_id'), row.get('name'), row.get('date')),
        'ActivityParticipants': lambda row: activity_key(row.get('school_id'), row.get('activity_name'), row.get('activity_date')) + '|' + normalized_key(row.get('username')),
        'LeadershipAwards': lambda row: award_key(row.get('recipient_username'), row.get('award_title'), row.get('year_awarded'), row.get('awarding_body')),
    }
    for sheet, rows in parsed.items():
        keys = [key_functions[sheet](row) for _, row in rows]
        duplicate_keys[sheet] = {key for key, count in Counter(keys).items() if key and count > 1}

    for row_number, raw in parsed['Schools']:
        row = RowPlan('Schools', row_number, raw)
        try:
            row.key = school_key(raw)
            if row.key in duplicate_keys['Schools']:
                raise ImportValidationError('Duplicate school key in workbook.')
            if not normalize(raw.get('name')):
                raise ImportValidationError('name is required.')
            email = normalize(raw.get('email'))
            if email:
                validate_email(email)
            proposed = {
                'name': normalize(raw.get('name')),
                'school_id': normalize(raw.get('school_id')),
                'school_type': _choice(raw.get('school_type'), School.SchoolType.choices, 'school_type', False),
                'category': normalize(raw.get('category')),
                'address': normalize(raw.get('address')),
                'location': normalize(raw.get('location')) or 'Philippines',
                'district': normalize(raw.get('district')),
                'division': normalize(raw.get('division')),
                'province': normalize(raw.get('province')),
                'region': normalize(raw.get('region')),
                'email': email,
                'phone': normalize(raw.get('phone')),
                'website': normalize(raw.get('website')),
                'founded_year': normalize(raw.get('founded_year')),
                'is_active': parse_bool(raw.get('is_active')),
            }
            existing = _find_school(raw)
            row.changes = _changes(existing, proposed) if existing else proposed
            row.action = ImportRowResult.Action.UPDATE if existing and row.changes else (ImportRowResult.Action.UNCHANGED if existing else ImportRowResult.Action.CREATE)
            staged_schools[row.key] = {'raw': raw, 'proposed': proposed, 'existing': existing}
        except (ImportValidationError, ValidationError) as exc:
            row.errors.append(str(exc))
            row.action = ImportRowResult.Action.ERROR
        plans.append(row)

    school_rows_by_key = {
        school_key(raw): (row_number, raw)
        for row_number, raw in parsed['Schools']
    }

    for row_number, raw in parsed['Users']:
        row = RowPlan('Users', row_number, raw)
        try:
            row.key = normalized_key(raw.get('username'))
            if not row.key:
                raise ImportValidationError('username is required.')
            if row.key in duplicate_keys['Users']:
                raise ImportValidationError('Duplicate username in workbook.')
            role = _choice(raw.get('role'), User.Role.choices, 'role')
            person_type = _choice(raw.get('person_type'), Person.Type.choices, 'person_type')
            expected_role = ROLE_FOR_PERSON[person_type]
            if role != expected_role and role not in {User.Role.ADMIN, User.Role.ENCODER}:
                raise ImportValidationError(f'role {role} is incompatible with person_type {person_type}; expected {expected_role}.')
            school_identifier = normalized_key(raw.get('school_id'))
            if school_identifier and school_identifier not in staged_schools:
                matches = School.objects.filter(school_id__iexact=normalize(raw.get('school_id')))
                if matches.count() != 1:
                    raise ImportValidationError('school_id must match exactly one database or workbook school.')
            email = normalize(raw.get('email'))
            if email:
                validate_email(email)
            existing = User.objects.filter(username__iexact=normalize(raw.get('username')))
            if existing.count() > 1:
                raise ImportValidationError('Username matches multiple accounts case-insensitively.')
            user = existing.first()
            user_values = {
                'username': normalize(raw.get('username')),
                'first_name': normalize(raw.get('first_name')),
                'last_name': normalize(raw.get('last_name')),
                'email': email,
                'role': role,
                'is_active': parse_bool(raw.get('is_active')),
            }
            person_values = {
                'type': person_type,
                'contact_number': normalize(raw.get('contact_number')),
                'address': normalize(raw.get('address')),
                'bio': normalize(raw.get('bio')),
                'student_id': normalize(raw.get('student_id')),
                'year_level': normalize(raw.get('year_level')),
                'course_program': normalize(raw.get('course_program')),
                'section': normalize(raw.get('section')),
                'scholarship_type': normalize(raw.get('scholarship_type')),
                'year_started': normalize(raw.get('year_started')),
                'year_ended': normalize(raw.get('year_ended')),
            }
            changes = _changes(user, user_values) if user else user_values.copy()
            person = Person.objects.filter(user=user).first() if user else None
            if person:
                for name, change in _changes(person, person_values).items():
                    changes[f'person.{name}'] = change
            elif user:
                changes['person'] = {'from': None, 'to': 'create profile'}
            row.changes = changes
            row.action = ImportRowResult.Action.UPDATE if user and changes else (ImportRowResult.Action.UNCHANGED if user else ImportRowResult.Action.CREATE)
            staged_users[row.key] = {'raw': raw, 'user_values': user_values, 'person_values': person_values, 'existing': user}
        except (ImportValidationError, ValidationError) as exc:
            row.errors.append(str(exc))
            row.action = ImportRowResult.Action.ERROR
        plans.append(row)

    for plan_row in [item for item in plans if item.sheet == 'Schools' and not item.errors]:
        _, raw = school_rows_by_key[plan_row.key]
        principal_username = normalized_key(raw.get('principal_username'))
        if not principal_username:
            continue
        staged_user = staged_users.get(principal_username)
        principal = staged_user['existing'] if staged_user else User.objects.filter(
            username__iexact=normalize(raw.get('principal_username'))
        ).first()
        if staged_user:
            role = staged_user['user_values']['role']
            principal_school_id = normalized_key(staged_user['raw'].get('school_id'))
        elif principal and hasattr(principal, 'person'):
            role = principal.role
            principal_school_id = normalized_key(
                principal.person.school.school_id if principal.person.school else ''
            )
        else:
            role = None
            principal_school_id = ''
        school_identifier = normalized_key(raw.get('school_id'))
        if role != User.Role.PRINCIPAL or not school_identifier or principal_school_id != school_identifier:
            plan_row.errors.append(
                'principal_username must identify a PRINCIPAL profile assigned to this school_id.'
            )
            plan_row.action = ImportRowResult.Action.ERROR

    for row_number, raw in parsed['Activities']:
        row = RowPlan('Activities', row_number, raw)
        try:
            activity_date = parse_date(raw.get('date'), 'date')
            row.key = activity_key(raw.get('school_id'), raw.get('name'), activity_date)
            if row.key in duplicate_keys['Activities']:
                raise ImportValidationError('Duplicate activity key in workbook.')
            school_identifier = normalized_key(raw.get('school_id'))
            staged_school = staged_schools.get(school_identifier)
            school = staged_school['existing'] if staged_school else None
            if not school:
                matches = School.objects.filter(school_id__iexact=normalize(raw.get('school_id')))
                if matches.count() > 1:
                    raise ImportValidationError('school_id matches multiple database schools.')
                school = matches.first()
            if not staged_school and not school:
                raise ImportValidationError('school_id does not match a database or workbook school.')
            proposed = {
                'name': normalize(raw.get('name')),
                'date': activity_date,
                'description': normalize(raw.get('description')),
                'is_approved': parse_bool(raw.get('is_approved'), False),
            }
            if not proposed['name']:
                raise ImportValidationError('name is required.')
            existing = Activity.objects.filter(school=school, name__iexact=proposed['name'], date=activity_date) if school else Activity.objects.none()
            if existing.count() > 1:
                raise ImportValidationError('Activity key matches multiple database records.')
            activity = existing.first()
            row.changes = _changes(activity, proposed) if activity else proposed
            row.action = ImportRowResult.Action.UPDATE if activity and row.changes else (ImportRowResult.Action.UNCHANGED if activity else ImportRowResult.Action.CREATE)
            staged_activities[row.key] = {'raw': raw, 'proposed': proposed, 'existing': activity}
        except ImportValidationError as exc:
            row.errors.append(str(exc))
            row.action = ImportRowResult.Action.ERROR
        plans.append(row)

    for row_number, raw in parsed['ActivityParticipants']:
        row = RowPlan('ActivityParticipants', row_number, raw)
        try:
            activity_date = parse_date(raw.get('activity_date'), 'activity_date')
            base_key = activity_key(raw.get('school_id'), raw.get('activity_name'), activity_date)
            row.key = base_key + '|' + normalized_key(raw.get('username'))
            if row.key in duplicate_keys['ActivityParticipants']:
                raise ImportValidationError('Duplicate participant assignment in workbook.')
            activity_stage = staged_activities.get(base_key)
            username = normalized_key(raw.get('username'))
            user_stage = staged_users.get(username)
            activity = activity_stage['existing'] if activity_stage else None
            user = user_stage['existing'] if user_stage else User.objects.filter(username__iexact=normalize(raw.get('username'))).first()
            person = Person.objects.filter(user=user).first() if user else None
            if not activity_stage and not activity:
                raise ImportValidationError('Activity does not match an Activities row or existing activity.')
            if not user_stage and not person:
                raise ImportValidationError('username does not match a workbook or existing Person profile.')
            exists = bool(activity and person and activity.participants.filter(pk=person.pk).exists())
            row.action = ImportRowResult.Action.UNCHANGED if exists else ImportRowResult.Action.CREATE
            row.changes = {} if exists else {'participant': {'from': None, 'to': normalize(raw.get('username'))}}
        except ImportValidationError as exc:
            row.errors.append(str(exc))
            row.action = ImportRowResult.Action.ERROR
        plans.append(row)

    for row_number, raw in parsed['LeadershipAwards']:
        row = RowPlan('LeadershipAwards', row_number, raw)
        try:
            year = parse_year(raw.get('year_awarded'), 'year_awarded')
            row.key = award_key(raw.get('recipient_username'), raw.get('award_title'), year, raw.get('awarding_body'))
            if row.key in duplicate_keys['LeadershipAwards']:
                raise ImportValidationError('Duplicate leadership award key in workbook.')
            username = normalized_key(raw.get('recipient_username'))
            user_stage = staged_users.get(username)
            user = user_stage['existing'] if user_stage else User.objects.filter(username__iexact=normalize(raw.get('recipient_username'))).first()
            person = Person.objects.filter(user=user).first() if user else None
            if not user_stage and not person:
                raise ImportValidationError('recipient_username does not match a workbook or existing Person profile.')
            level = _choice(raw.get('award_level'), LeadershipAward.AwardLevel.choices, 'award_level')
            title = normalize(raw.get('award_title'))
            body = normalize(raw.get('awarding_body'))
            existing = LeadershipAward.objects.filter(recipient=person, award_title__iexact=title, year_awarded=year, awarding_body__iexact=body) if person else LeadershipAward.objects.none()
            if existing.count() > 1:
                raise ImportValidationError('Award key matches multiple database records.')
            award = existing.first()
            proposed = {'award_title': title, 'award_level': level, 'year_awarded': year, 'awarding_body': body, 'description': normalize(raw.get('description'))}
            row.changes = _changes(award, proposed) if award else proposed
            row.action = ImportRowResult.Action.UPDATE if award and row.changes else (ImportRowResult.Action.UNCHANGED if award else ImportRowResult.Action.CREATE)
        except ImportValidationError as exc:
            row.errors.append(str(exc))
            row.action = ImportRowResult.Action.ERROR
        plans.append(row)

    return ImportPlan(plans)


def save_preview(job: ImportJob) -> ImportPlan:
    plan = build_plan(bytes(job.workbook_data))
    job.row_results.all().delete()
    ImportRowResult.objects.bulk_create([
        ImportRowResult(
            job=job,
            sheet_name=row.sheet,
            row_number=row.row_number,
            external_key=row.key,
            action=row.action,
            source_data={key: safe_json_value(value) for key, value in row.source.items()},
            changes=json_safe(row.changes),
            errors=row.errors,
        ) for row in plan.rows
    ])
    job.total_rows = len(plan.rows)
    job.create_count = plan.count(ImportRowResult.Action.CREATE)
    job.update_count = plan.count(ImportRowResult.Action.UPDATE)
    job.unchanged_count = plan.count(ImportRowResult.Action.UNCHANGED)
    job.error_count = plan.count(ImportRowResult.Action.ERROR)
    job.status = ImportJob.Status.INVALID if job.error_count else ImportJob.Status.READY
    job.previewed_at = timezone.now()
    job.failure_message = ''
    job.save()
    return plan


def _plan_matches_saved_preview(job: ImportJob, plan: ImportPlan) -> bool:
    saved = {
        (row.sheet_name, row.row_number): (row.action, row.changes, row.errors)
        for row in job.row_results.all()
    }
    current = {
        (row.sheet, row.row_number): (row.action, json_safe(row.changes), row.errors)
        for row in plan.rows
    }
    return saved == current


def create_job(upload, user) -> ImportJob:
    content = upload.read()
    job = ImportJob.objects.create(
        uploaded_by=user,
        original_filename=normalize(upload.name)[:255],
        checksum=hashlib.sha256(content).hexdigest(),
        workbook_data=content,
    )
    try:
        save_preview(job)
    except Exception as exc:
        job.status = ImportJob.Status.INVALID
        job.error_count = 1
        job.failure_message = str(exc)
        job.previewed_at = timezone.now()
        job.save()
    return job


def _school_from_identifier(identifier, school_map):
    key = normalized_key(identifier)
    if not key:
        return None
    if key in school_map:
        return school_map[key]
    matches = School.objects.filter(school_id__iexact=normalize(identifier))
    if matches.count() != 1:
        raise ImportValidationError(f'Unable to resolve school_id {identifier!r} during apply.')
    return matches.first()


def apply_job(job: ImportJob) -> None:
    if not job.can_apply:
        raise ImportValidationError('Only a valid preview can be applied.')
    successful_duplicate = ImportJob.objects.filter(
        checksum=job.checksum, status=ImportJob.Status.APPLIED
    ).exclude(pk=job.pk).exists()
    if successful_duplicate:
        raise ImportValidationError('This exact workbook has already been applied.')

    current_plan = build_plan(bytes(job.workbook_data))
    if current_plan.errors or not _plan_matches_saved_preview(job, current_plan):
        save_preview(job)
        raise ImportValidationError('Data changed after preview. Review the refreshed validation results.')

    parsed = read_workbook(bytes(job.workbook_data))
    try:
        with transaction.atomic():
            job.status = ImportJob.Status.APPLYING
            job.save(update_fields=['status', 'updated_at'])
            school_map = {}
            for _, row in parsed['Schools']:
                existing = _find_school(row)
                values = {
                    'name': normalize(row.get('name')), 'school_id': normalize(row.get('school_id')),
                    'school_type': _choice(row.get('school_type'), School.SchoolType.choices, 'school_type', False),
                    'category': normalize(row.get('category')), 'address': normalize(row.get('address')),
                    'location': normalize(row.get('location')) or 'Philippines', 'district': normalize(row.get('district')),
                    'division': normalize(row.get('division')), 'province': normalize(row.get('province')),
                    'region': normalize(row.get('region')), 'email': normalize(row.get('email')),
                    'phone': normalize(row.get('phone')), 'website': normalize(row.get('website')),
                    'founded_year': normalize(row.get('founded_year')), 'is_active': parse_bool(row.get('is_active')),
                }
                school = existing or School()
                for name, value in values.items():
                    setattr(school, name, value)
                school.save()
                school_map[school_key(row)] = school

            user_map = {}
            for _, row in parsed['Users']:
                username = normalize(row.get('username'))
                user = User.objects.select_for_update().filter(username__iexact=username).first() or User(username=username)
                user.first_name = normalize(row.get('first_name'))
                user.last_name = normalize(row.get('last_name'))
                user.email = normalize(row.get('email'))
                user.role = _choice(row.get('role'), User.Role.choices, 'role')
                user.is_active = parse_bool(row.get('is_active'))
                if not user.pk:
                    user.set_unusable_password()
                    user.must_change_password = True
                user.save()
                school = _school_from_identifier(row.get('school_id'), school_map)
                person, _ = Person.objects.get_or_create(user=user)
                for name in ['contact_number', 'address', 'bio', 'student_id', 'year_level', 'course_program', 'section', 'scholarship_type', 'year_started', 'year_ended']:
                    setattr(person, name, normalize(row.get(name)))
                person.type = _choice(row.get('person_type'), Person.Type.choices, 'person_type')
                person.school = school
                person.save()
                user_map[normalized_key(username)] = user

            for _, row in parsed['Schools']:
                principal_username = normalize(row.get('principal_username'))
                if principal_username:
                    school = school_map[school_key(row)]
                    principal = user_map.get(normalized_key(principal_username)) or User.objects.filter(username__iexact=principal_username).first()
                    if not principal or principal.role != User.Role.PRINCIPAL or not hasattr(principal, 'person') or principal.person.school_id != school.pk:
                        raise ImportValidationError(f'Principal {principal_username!r} is not a principal profile in {school.name}.')
                    if school.principal_id != principal.pk:
                        school.principal = principal
                        school.save()

            activity_map = {}
            for _, row in parsed['Activities']:
                school = _school_from_identifier(row.get('school_id'), school_map)
                activity_date = parse_date(row.get('date'), 'date')
                activity = Activity.objects.filter(school=school, name__iexact=normalize(row.get('name')), date=activity_date).first() or Activity(school=school)
                activity.name = normalize(row.get('name'))
                activity.date = activity_date
                activity.description = normalize(row.get('description'))
                activity.is_approved = parse_bool(row.get('is_approved'), False)
                activity.approved_by = job.uploaded_by if activity.is_approved else None
                activity.save()
                activity_map[activity_key(row.get('school_id'), row.get('name'), activity_date)] = activity

            for _, row in parsed['ActivityParticipants']:
                key = activity_key(row.get('school_id'), row.get('activity_name'), parse_date(row.get('activity_date'), 'activity_date'))
                activity = activity_map.get(key)
                if not activity:
                    school = _school_from_identifier(row.get('school_id'), school_map)
                    activity = Activity.objects.get(school=school, name__iexact=normalize(row.get('activity_name')), date=parse_date(row.get('activity_date'), 'activity_date'))
                user = user_map.get(normalized_key(row.get('username'))) or User.objects.get(username__iexact=normalize(row.get('username')))
                activity.participants.add(user.person)

            for _, row in parsed['LeadershipAwards']:
                user = user_map.get(normalized_key(row.get('recipient_username'))) or User.objects.get(username__iexact=normalize(row.get('recipient_username')))
                person = user.person
                year = parse_year(row.get('year_awarded'), 'year_awarded')
                title = normalize(row.get('award_title'))
                body = normalize(row.get('awarding_body'))
                award = LeadershipAward.objects.filter(recipient=person, award_title__iexact=title, year_awarded=year, awarding_body__iexact=body).first() or LeadershipAward(recipient=person)
                award.award_title = title
                award.award_level = _choice(row.get('award_level'), LeadershipAward.AwardLevel.choices, 'award_level')
                award.year_awarded = year
                award.awarding_body = body
                award.description = normalize(row.get('description'))
                award.school = _school_from_identifier(row.get('school_id'), school_map) or person.school
                award.save()

            job.row_results.update(applied=True)
            job.status = ImportJob.Status.APPLIED
            job.applied_at = timezone.now()
            job.failure_message = ''
            job.save()
    except Exception as exc:
        job.status = ImportJob.Status.FAILED
        job.failure_message = str(exc)
        job.save()
        raise
